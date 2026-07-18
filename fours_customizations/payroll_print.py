"""
payroll_print.py — "Print Payroll" button on Payroll Entry.

The button shows in every document state (draft, submitted, cancelled) as long
as the entry has salary slips, and offers four outputs — each as an inline
(printable) landscape PDF or an .xlsx download:

  • Payroll          — one row per employee: earnings, the attendance
                       deductions, other/total deductions, net pay and the
                       violation-day / overtime-hour counts, with a totals row.
  • Attendance Logs  — a monthly punch grid (IN over OUT per day) for every
                       employee of the entry that has a Shift Assignment in the
                       period; mirrors the physical "Attend. Logs" sheet.

All outputs are styled with the 4S Industries brand colours taken from the logo
on multax.kit.africa: green #8FC643 on black #221E1F.

Payroll columns:
  Employee, Designation, Basic, Sales Commission, Overtime Pay,
  Total Earnings, Late Deduction, No Checkout Deduction,
  Early Exit Deduction, Absent Deduction, Other Deductions,
  Total Deductions, Net Pay, Days Absent, Early Exit Days,
  No Checkout Days, Overtime Hours
"""

from __future__ import annotations

import base64
import mimetypes
import os
from datetime import timedelta
from io import BytesIO

import frappe
from frappe.utils import flt, get_datetime, getdate

from fours_customizations.fours_customizations.doctype.four_s_industries_settings.four_s_industries_settings import (
	get_setting,
)

BRAND_GREEN = "8FC643"
BRAND_BLACK = "221E1F"
BRAND_GREEN_TINT = "F1F8E3"  # light tint for zebra rows

# Overtime earning. The active structures use "Designation Overtime Pay"; the
# manual "Late Exit (Over Time)" earning is treated as overtime too so it shows
# in the Overtime Pay column rather than vanishing into the gross total only.
_OVERTIME_COMPONENTS = ("Designation Overtime Pay", "Late Exit (Over Time)")

# Each attendance-deduction column → the salary components that feed it. The
# active "4s Salary Structure with Deductions" (and the current slips) use the
# canonical "* Deduction" names; the legacy variants (Missed Checkout, Early
# Exit, Absence) are aliased so an older or alternate structure never leaks an
# attendance deduction into "Other Deductions". Alias groups are disjoint, so
# nothing is double-counted. Verified against multax.kit.africa: the June 2026
# run uses Absent / Late / Early Exit / No Checkout Deduction.
_ATTENDANCE_DEDUCTIONS = {
	"late_deduction": ("Late Deduction",),
	"no_checkout_deduction": ("No Checkout Deduction", "Missed Checkout"),
	"early_exit_deduction": ("Early Exit Deduction", "Early Exit"),
	"absent_deduction": ("Absent Deduction", "Absence"),
}

# (row key, column label, kind) — kind drives alignment & number format.
COLUMNS = [
	("employee", "Employee", "text"),
	("designation", "Designation", "text"),
	("basic", "Basic", "currency"),
	("sales_commission", "Sales Commission", "currency"),
	("overtime_pay", "Overtime Pay", "currency"),
	("total_earnings", "Total Earnings", "currency"),
	("late_deduction", "Late Deduction", "currency"),
	("no_checkout_deduction", "No Checkout Deduction", "currency"),
	("early_exit_deduction", "Early Exit Deduction", "currency"),
	("absent_deduction", "Absent Deduction", "currency"),
	("other_deductions", "Other Deductions", "currency"),
	("total_deductions", "Total Deductions", "currency"),
	("net_pay", "Net Pay", "currency"),
	("days_absent", "Days Absent", "int"),
	("early_exit_days", "Early Exit Days", "int"),
	("no_checkout_days", "No Checkout Days", "int"),
	("overtime_hours", "Overtime Hours", "hours"),
]


@frappe.whitelist()
def get_print_allowed(payroll_entry: str) -> dict:
	"""The form script asks this on refresh: the button shows in every state of
	the Payroll Entry (draft, submitted, cancelled) as long as it has at least
	one salary slip to print."""
	frappe.has_permission("Payroll Entry", "read", payroll_entry, throw=True)
	total = frappe.db.count("Salary Slip", {"payroll_entry": payroll_entry})
	return {"allowed": bool(total), "total": total}


@frappe.whitelist()
def download_payroll_excel(payroll_entry: str):
	pe = frappe.get_doc("Payroll Entry", payroll_entry)
	pe.check_permission("read")
	rows, totals = _get_payroll_rows(pe)

	frappe.local.response.filename = f"Payroll {pe.start_date} to {pe.end_date}.xlsx"
	frappe.local.response.filecontent = _build_excel(pe, rows, totals)
	frappe.local.response.type = "binary"


@frappe.whitelist()
def download_payroll_pdf(payroll_entry: str):
	from frappe.utils.pdf import get_pdf

	pe = frappe.get_doc("Payroll Entry", payroll_entry)
	pe.check_permission("read")
	rows, totals = _get_payroll_rows(pe)

	# type "pdf" serves the file inline, so the browser's print dialog works
	# straight from the new tab.
	frappe.local.response.filename = f"Payroll {pe.start_date} to {pe.end_date}.pdf"
	frappe.local.response.filecontent = get_pdf(_build_pdf_html(pe, rows, totals), {"orientation": "Landscape"})
	frappe.local.response.type = "pdf"


@frappe.whitelist()
def download_attendance_logs_excel(payroll_entry: str):
	pe = frappe.get_doc("Payroll Entry", payroll_entry)
	pe.check_permission("read")
	data = _attendance_log_data(pe)

	frappe.local.response.filename = f"Attendance Logs {pe.start_date} to {pe.end_date}.xlsx"
	frappe.local.response.filecontent = _build_attendance_excel(pe, data)
	frappe.local.response.type = "binary"


@frappe.whitelist()
def download_attendance_logs_pdf(payroll_entry: str):
	from frappe.utils.pdf import get_pdf

	pe = frappe.get_doc("Payroll Entry", payroll_entry)
	pe.check_permission("read")
	data = _attendance_log_data(pe)

	frappe.local.response.filename = f"Attendance Logs {pe.start_date} to {pe.end_date}.pdf"
	frappe.local.response.filecontent = get_pdf(
		_build_attendance_pdf_html(pe, data), {"orientation": "Landscape"}
	)
	frappe.local.response.type = "pdf"


# ── dataset ─────────────────────────────────────────────────────────────────

def _get_payroll_rows(pe) -> tuple[list[dict], dict]:
	"""One row per employee of the Payroll Entry, plus a totals dict.

	The printout is available in every document state, so slips in any
	docstatus qualify — draft (0), submitted (1) and cancelled (2); when an
	employee has several (e.g. amended), the most recently created wins.
	"""
	slips = _latest_slips(pe.name)
	if not slips:
		frappe.throw("No salary slips found for this Payroll Entry.")

	components = _component_amounts([s.name for s in slips])
	commission_component = get_setting("commission_salary_component", "Sales Commission")

	rows = []
	for slip in slips:
		earnings = components.get(slip.name, {}).get("earnings", {})
		deductions = components.get(slip.name, {}).get("deductions", {})

		attendance = {
			key: sum(flt(deductions.get(name)) for name in names)
			for key, names in _ATTENDANCE_DEDUCTIONS.items()
		}
		# Everything on the slip that isn't one of the aliased attendance
		# deductions (loans, advances, shortages, …) is "Other Deductions".
		other_deductions = flt(slip.total_deduction) - sum(attendance.values())
		counts = _attendance_counts(slip.employee, slip.start_date, slip.end_date, pe.company)

		rows.append({
			"employee": slip.employee_name or slip.employee,
			"designation": slip.designation or "",
			"basic": _basic_amount(slip, earnings),
			"sales_commission": flt(earnings.get(commission_component)),
			"overtime_pay": sum(flt(earnings.get(c)) for c in _OVERTIME_COMPONENTS),
			"total_earnings": flt(slip.gross_pay),
			"late_deduction": attendance["late_deduction"],
			"no_checkout_deduction": attendance["no_checkout_deduction"],
			"early_exit_deduction": attendance["early_exit_deduction"],
			"absent_deduction": attendance["absent_deduction"],
			"other_deductions": flt(other_deductions, 2),
			"total_deductions": flt(slip.total_deduction),
			"net_pay": flt(slip.net_pay),
			"days_absent": counts["absent"],
			"early_exit_days": counts["early_exit"],
			"no_checkout_days": counts["no_checkout"],
			"overtime_hours": counts["overtime_hours"],
		})

	totals = {key: sum(r[key] for r in rows) for key, _label, kind in COLUMNS if kind != "text"}
	return rows, totals


def _latest_slips(payroll_entry: str) -> list:
	slips = frappe.get_all(
		"Salary Slip",
		filters={"payroll_entry": payroll_entry, "docstatus": ["in", [0, 1, 2]]},
		fields=[
			"name", "employee", "employee_name", "designation",
			"start_date", "end_date",
			"gross_pay", "total_deduction", "net_pay", "custom_basic_pay",
		],
		order_by="creation asc",
	)
	latest = {}
	for slip in slips:  # ascending creation → the newest slip per employee wins
		latest[slip.employee] = slip
	return sorted(latest.values(), key=lambda s: (s.employee_name or s.employee or ""))


def _component_amounts(slip_names: list[str]) -> dict:
	"""{slip name: {"earnings": {component: amount}, "deductions": {...}}}"""
	out: dict = {}
	if not slip_names:
		return out
	details = frappe.get_all(
		"Salary Detail",
		filters={"parent": ["in", slip_names], "parentfield": ["in", ["earnings", "deductions"]]},
		fields=["parent", "parentfield", "salary_component", "amount"],
	)
	for d in details:
		table = out.setdefault(d.parent, {"earnings": {}, "deductions": {}})[d.parentfield]
		table[d.salary_component] = table.get(d.salary_component, 0.0) + flt(d.amount)
	return out


def _basic_amount(slip, earnings: dict) -> float:
	"""The paid Basic earning; falls back to the stamped base when the slip's
	structure names its base component something other than "Basic"."""
	if "Basic" in earnings:
		return flt(earnings["Basic"])
	for component, amount in earnings.items():
		if "basic" in (component or "").lower():
			return flt(amount)
	return flt(slip.custom_basic_pay)


def _attendance_counts(employee, start_date, end_date, company) -> dict:
	"""Violation-day counts and overtime hours for the slip period, using the
	same shift-assignment + holiday gating the deductions themselves use."""
	counts = {"absent": 0, "early_exit": 0, "no_checkout": 0, "overtime_hours": 0.0}
	try:
		from fours_customizations.overtime_utils import calculate_designation_overtime
		from fours_customizations.salary_slip_handler import _holiday_dates, _shift_assigned_dates

		start, end = getdate(start_date), getdate(end_date)
		eligible = _shift_assigned_dates(employee, start, end) - _holiday_dates(
			employee, company, start, end
		)
		if eligible:
			records = frappe.get_all(
				"Attendance",
				filters={
					"employee": employee,
					"attendance_date": ["between", [start, end]],
					"docstatus": 1,
				},
				fields=["status", "attendance_date", "out_time", "early_exit"],
			)
			for att in records:
				if getdate(att.attendance_date) not in eligible:
					continue
				if att.status == "Absent":
					counts["absent"] += 1
				if att.early_exit == 1:
					counts["early_exit"] += 1
				if att.status in ("Present", "Half Day") and not att.out_time:
					counts["no_checkout"] += 1

		counts["overtime_hours"] = flt(
			calculate_designation_overtime(employee, start_date, end_date).get("total_hours")
		)
	except Exception:
		frappe.log_error(frappe.get_traceback(), "4S Payroll Print: attendance counts failed")
	return counts


# ── Excel ───────────────────────────────────────────────────────────────────

_NUMBER_FORMATS = {"currency": "#,##0.00", "int": "0", "hours": "0.00"}


def _build_excel(pe, rows: list[dict], totals: dict) -> bytes:
	import openpyxl
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Payroll"
	n_cols = len(COLUMNS)

	green_fill = PatternFill("solid", fgColor=BRAND_GREEN)
	black_fill = PatternFill("solid", fgColor=BRAND_BLACK)
	tint_fill = PatternFill("solid", fgColor=BRAND_GREEN_TINT)
	thin = Side(style="thin", color="D9D9D9")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
	title = ws.cell(row=1, column=1, value=pe.company)
	title.font = Font(bold=True, size=16, color="FFFFFF")
	title.fill = black_fill
	title.alignment = Alignment(horizontal="center", vertical="center")
	ws.row_dimensions[1].height = 28
	for col in range(2, n_cols + 1):
		ws.cell(row=1, column=col).fill = black_fill

	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
	subtitle = ws.cell(
		row=2, column=1, value=f"Payroll — {pe.start_date} to {pe.end_date}  ({pe.name})"
	)
	subtitle.font = Font(bold=True, size=11, color=BRAND_BLACK)
	subtitle.fill = green_fill
	subtitle.alignment = Alignment(horizontal="center", vertical="center")
	ws.row_dimensions[2].height = 20
	for col in range(2, n_cols + 1):
		ws.cell(row=2, column=col).fill = green_fill

	header_row = 4
	for idx, (_key, label, _kind) in enumerate(COLUMNS, start=1):
		cell = ws.cell(row=header_row, column=idx, value=label)
		cell.font = Font(bold=True, color=BRAND_BLACK, size=9)
		cell.fill = green_fill
		cell.border = border
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
	ws.row_dimensions[header_row].height = 30
	ws.freeze_panes = f"A{header_row + 1}"

	for r_idx, row in enumerate(rows, start=header_row + 1):
		for c_idx, (key, _label, kind) in enumerate(COLUMNS, start=1):
			cell = ws.cell(row=r_idx, column=c_idx, value=row[key])
			cell.border = border
			cell.font = Font(size=9, color=BRAND_BLACK)
			if (r_idx - header_row) % 2 == 0:
				cell.fill = tint_fill
			if kind in _NUMBER_FORMATS:
				cell.number_format = _NUMBER_FORMATS[kind]
				cell.alignment = Alignment(horizontal="right")

	total_row = header_row + len(rows) + 1
	for c_idx, (key, _label, kind) in enumerate(COLUMNS, start=1):
		value = "TOTAL" if c_idx == 1 else totals.get(key, "")
		cell = ws.cell(row=total_row, column=c_idx, value=value)
		cell.font = Font(bold=True, color="FFFFFF", size=9)
		cell.fill = black_fill
		if kind in _NUMBER_FORMATS and value != "":
			cell.number_format = _NUMBER_FORMATS[kind]
			cell.alignment = Alignment(horizontal="right")

	for c_idx, (_key, label, kind) in enumerate(COLUMNS, start=1):
		width = 22 if kind == "text" else max(len(label) + 2, 12)
		ws.column_dimensions[get_column_letter(c_idx)].width = width

	buf = BytesIO()
	wb.save(buf)
	return buf.getvalue()


# ── PDF ─────────────────────────────────────────────────────────────────────

def _fmt(value, kind: str) -> str:
	if kind == "currency":
		return f"{flt(value):,.2f}"
	if kind == "int":
		return f"{int(value or 0)}"
	if kind == "hours":
		return f"{flt(value):,.2f}"
	return frappe.utils.escape_html(str(value or ""))


def _logo_data_uri() -> str | None:
	"""Company/site logo embedded as a data URI (wkhtmltopdf can't rely on
	fetching site URLs). Returns None when no local logo file is found."""
	candidates = [
		frappe.db.get_single_value("Website Settings", "app_logo"),
		frappe.db.get_single_value("Website Settings", "banner_image"),
		"/files/logo.png",
	]
	for logo in candidates:
		if not logo or logo.startswith("http"):
			continue
		relative = logo.lstrip("/")
		path = (
			frappe.get_site_path(relative)
			if relative.startswith("private/")
			else frappe.get_site_path("public", relative)
		)
		if os.path.exists(path):
			mime = mimetypes.guess_type(path)[0] or "image/png"
			with open(path, "rb") as f:
				return f"data:{mime};base64,{base64.b64encode(f.read()).decode()}"
	return None


def _build_pdf_html(pe, rows: list[dict], totals: dict) -> str:
	logo = _logo_data_uri()
	logo_html = f'<img src="{logo}" style="height:52px;">' if logo else ""

	header_cells = "".join(f"<th>{label}</th>" for _key, label, _kind in COLUMNS)

	body_rows = []
	for row in rows:
		cells = "".join(
			f'<td class="{ "num" if kind != "text" else "" }">{_fmt(row[key], kind)}</td>'
			for key, _label, kind in COLUMNS
		)
		body_rows.append(f"<tr>{cells}</tr>")

	total_cells = ['<td class="label">TOTAL</td>']
	for key, _label, kind in COLUMNS[1:]:
		total_cells.append(
			f'<td class="num">{_fmt(totals[key], kind)}</td>' if kind != "text" else "<td></td>"
		)

	return f"""
<style>
	* {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
	body {{ font-family: "Helvetica Neue", Helvetica, Arial, sans-serif; color: #{BRAND_BLACK}; margin: 0; }}
	.band {{ border-bottom: 4px solid #{BRAND_GREEN}; padding: 6px 0 10px 0; }}
	.band table {{ width: 100%; border: none; }}
	.band td {{ border: none; vertical-align: middle; }}
	.company {{ font-size: 18px; font-weight: bold; }}
	.meta {{ font-size: 10px; color: #555; }}
	.title {{ font-size: 13px; font-weight: bold; color: #{BRAND_GREEN}; text-align: right;
		text-transform: uppercase; letter-spacing: 2px; }}
	.period {{ font-size: 10px; text-align: right; color: #{BRAND_BLACK}; }}
	table.payroll {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 8px; }}
	table.payroll th {{ background: #{BRAND_GREEN}; color: #{BRAND_BLACK}; font-weight: bold;
		padding: 5px 4px; border: 1px solid #{BRAND_GREEN}; text-align: center; }}
	table.payroll td {{ border: 1px solid #DDDDDD; padding: 4px; }}
	table.payroll tbody tr:nth-child(even) td {{ background: #{BRAND_GREEN_TINT}; }}
	td.num {{ text-align: right; white-space: nowrap; }}
	/* Totals live in <tfoot> so the zebra rule on tbody can never override the
	   black band (a specificity clash that once made this row invisible). */
	table.payroll tfoot td {{ background: #{BRAND_BLACK}; color: #FFFFFF; font-weight: bold;
		border: 1px solid #{BRAND_BLACK}; text-align: right; white-space: nowrap; }}
	table.payroll tfoot td.label {{ text-align: left; letter-spacing: 1px; }}
</style>
<div class="band">
	<table>
		<tr>
			<td style="width:60px;">{logo_html}</td>
			<td>
				<div class="company">{frappe.utils.escape_html(pe.company or "")}</div>
				<div class="meta">{frappe.utils.escape_html(pe.name)}</div>
			</td>
			<td>
				<div class="title">Payroll</div>
				<div class="period">{pe.start_date} to {pe.end_date}</div>
			</td>
		</tr>
	</table>
</div>
<table class="payroll">
	<thead><tr>{header_cells}</tr></thead>
	<tbody>
		{"".join(body_rows)}
	</tbody>
	<tfoot><tr>{"".join(total_cells)}</tr></tfoot>
</table>
"""


# ═════════════════════════════════════════════════════════════════════════════
# Attendance Logs — a monthly punch-log grid (IN over OUT per day) for every
# employee of the Payroll Entry who has a Shift Assignment in the period. PDF
# and Excel, same 4S branding. Mirrors the physical "Attend. Logs" sheet.
# ═════════════════════════════════════════════════════════════════════════════

# Days per strip in the PDF grid — keeps a landscape A4 row readable by
# splitting a ~30-day month into two half-month strips, as on the paper form.
_LOG_CHUNK = 16


def _hhmm(dt) -> str:
	"""'HH:MM' from an Attendance in/out datetime; '' when unset."""
	if not dt:
		return ""
	try:
		return get_datetime(dt).strftime("%H:%M")
	except Exception:
		s = str(dt)
		return s[11:16] if len(s) >= 16 else ""


def _attendance_log_data(pe) -> dict:
	"""Grid data: the period's dates, the shift-assigned employees on this
	entry, and a {(employee, iso-date): punch} map from submitted Attendance."""
	start, end = getdate(pe.start_date), getdate(pe.end_date)
	period_dates = []
	day = start
	while day <= end:
		period_dates.append(day)
		day += timedelta(days=1)

	entry_employees = _entry_employees(pe.name)
	if not entry_employees:
		frappe.throw("No employees found for this Payroll Entry.")

	shift_emps = _shift_assigned_employees(entry_employees, start, end)
	if not shift_emps:
		frappe.throw("No employees on this Payroll Entry have a shift assigned in this period.")

	details = frappe.get_all(
		"Employee",
		filters={"name": ["in", list(shift_emps)]},
		fields=["name", "employee_name", "designation", "department", "employee_number"],
	)
	employees = sorted(
		(
			{
				"emp": e.name,
				"name": e.employee_name or e.name,
				"designation": e.designation or "",
				"department": e.department or "",
				"code": e.employee_number or e.name,
			}
			for e in details
		),
		key=lambda x: (x["name"] or "").lower(),
	)

	records = frappe.get_all(
		"Attendance",
		filters={
			"employee": ["in", list(shift_emps)],
			"attendance_date": ["between", [start, end]],
			"docstatus": 1,
		},
		fields=[
			"employee", "attendance_date", "in_time", "out_time",
			"status", "late_entry", "early_exit",
		],
	)
	punches: dict = {}
	for a in records:
		# Last write wins if a day somehow has duplicates; Attendance is 1/day.
		punches[(a.employee, str(getdate(a.attendance_date)))] = {
			"in": _hhmm(a.in_time),
			"out": _hhmm(a.out_time),
			"status": a.status or "",
			"late": bool(a.late_entry),
			"early": bool(a.early_exit),
		}

	return {"period_dates": period_dates, "employees": employees, "punches": punches}


def _entry_employees(payroll_entry: str) -> list[str]:
	"""Distinct employees with a salary slip in this entry (any docstatus)."""
	rows = frappe.get_all(
		"Salary Slip",
		filters={"payroll_entry": payroll_entry, "docstatus": ["in", [0, 1, 2]]},
		fields=["employee"],
		group_by="employee",
	)
	return [r.employee for r in rows if r.employee]


def _shift_assigned_employees(employees: list[str], start, end) -> set:
	"""Subset of *employees* with a submitted Shift Assignment overlapping the
	period. A blank end_date is treated as ongoing."""
	rows = frappe.get_all(
		"Shift Assignment",
		filters={"employee": ["in", employees], "docstatus": 1, "start_date": ["<=", end]},
		fields=["employee", "end_date"],
	)
	result: set = set()
	for r in rows:
		if r.end_date and getdate(r.end_date) < start:
			continue
		result.add(r.employee)
	return result


# ── attendance PDF ───────────────────────────────────────────────────────────

def _esc(value) -> str:
	return frappe.utils.escape_html(str(value or ""))


def _punch_cell(punch) -> str:
	"""One day cell: IN over OUT (two lines), or A for absent, blank if no record."""
	if not punch:
		return '<td class="empty"></td>'
	if punch["status"] == "Absent":
		return '<td class="absent">A</td>'
	in_txt = punch["in"] or "&middot;"
	out_txt = punch["out"] or "&mdash;"
	in_cls = " late" if punch["late"] else ""
	out_cls = " early" if punch["early"] else ""
	return (
		f'<td><span class="pin{in_cls}">{in_txt}</span>'
		f'<span class="pout{out_cls}">{out_txt}</span></td>'
	)


def _build_attendance_pdf_html(pe, data: dict) -> str:
	logo = _logo_data_uri()
	logo_html = f'<img src="{logo}" style="height:52px;">' if logo else ""

	dates = data["period_dates"]
	chunks = [dates[i:i + _LOG_CHUNK] for i in range(0, len(dates), _LOG_CHUNK)]

	blocks = []
	for emp in data["employees"]:
		strips = []
		for chunk in chunks:
			day_cells = "".join(
				f'<th>{d.day}<span class="wd">{d.strftime("%a")[:2]}</span></th>' for d in chunk
			)
			punch_cells = "".join(
				_punch_cell(data["punches"].get((emp["emp"], str(d)))) for d in chunk
			)
			strips.append(
				'<table class="logs">'
				f'<tr class="drow"><th class="rl">Day</th>{day_cells}</tr>'
				f'<tr class="prow"><td class="rl">IN<br>OUT</td>{punch_cells}</tr>'
				"</table>"
			)
		sub = " &middot; ".join(x for x in (emp["designation"], emp["department"]) if x)
		blocks.append(
			'<div class="emp">'
			f'<div class="emp-head"><span class="eid">{_esc(emp["code"])}</span> '
			f'{_esc(emp["name"])} <span class="sub">{_esc(sub)}</span></div>'
			f'{"".join(strips)}</div>'
		)

	return f"""
<style>
	* {{ -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
	body {{ font-family: "Helvetica Neue", Helvetica, Arial, sans-serif; color: #{BRAND_BLACK}; margin: 0; }}
	.band {{ border-bottom: 4px solid #{BRAND_GREEN}; padding: 6px 0 10px 0; }}
	.band table {{ width: 100%; border: none; }}
	.band td {{ border: none; vertical-align: middle; }}
	.company {{ font-size: 18px; font-weight: bold; }}
	.meta {{ font-size: 10px; color: #555; }}
	.title {{ font-size: 13px; font-weight: bold; color: #{BRAND_GREEN}; text-align: right;
		text-transform: uppercase; letter-spacing: 2px; }}
	.period {{ font-size: 10px; text-align: right; }}
	.emp {{ page-break-inside: avoid; margin-top: 12px; }}
	.emp-head {{ background: #{BRAND_BLACK}; color: #FFFFFF; font-size: 10px; font-weight: bold;
		padding: 4px 8px; }}
	.emp-head .eid {{ background: #{BRAND_GREEN}; color: #{BRAND_BLACK}; padding: 1px 6px;
		border-radius: 3px; margin-right: 6px; }}
	.emp-head .sub {{ font-weight: normal; color: #CFE8A6; }}
	table.logs {{ width: 100%; border-collapse: collapse; margin-top: 3px; font-size: 7px;
		table-layout: fixed; }}
	table.logs th, table.logs td {{ border: 1px solid #CCCCCC; text-align: center; padding: 1px;
		overflow: hidden; }}
	table.logs tr.drow th {{ background: #{BRAND_GREEN}; color: #{BRAND_BLACK}; font-weight: bold; }}
	table.logs th .wd {{ display: block; font-weight: normal; font-size: 6px; color: #3a5a12; }}
	table.logs td.rl, table.logs th.rl {{ background: #{BRAND_BLACK}; color: #FFFFFF; width: 34px;
		font-weight: bold; }}
	table.logs td .pin, table.logs td .pout {{ display: block; line-height: 1.35; }}
	table.logs td .pout {{ color: #555; }}
	.pin.late {{ color: #C0392B; font-weight: bold; }}
	.pout.early {{ color: #B9770E; font-weight: bold; }}
	td.absent {{ background: #FBE4E4; color: #C0392B; font-weight: bold; }}
	td.empty {{ background: #F5F5F5; }}
	.legend {{ margin-top: 10px; font-size: 8px; color: #444; }}
	.legend b {{ color: #{BRAND_BLACK}; }}
	.legend .sw {{ display: inline-block; padding: 0 5px; margin: 0 3px; border-radius: 2px; }}
</style>
<div class="band">
	<table>
		<tr>
			<td style="width:60px;">{logo_html}</td>
			<td>
				<div class="company">{_esc(pe.company)}</div>
				<div class="meta">{_esc(pe.name)}</div>
			</td>
			<td>
				<div class="title">Attendance Logs</div>
				<div class="period">{pe.start_date} to {pe.end_date}</div>
			</td>
		</tr>
	</table>
</div>
{"".join(blocks)}
<div class="legend">
	Each day shows <b>IN</b> (top) over <b>OUT</b> (bottom).
	<span class="sw" style="background:#FBE4E4;color:#C0392B;">A</span> Absent &nbsp;
	<span style="color:#C0392B;font-weight:bold;">red IN</span> = late &nbsp;
	<span style="color:#B9770E;font-weight:bold;">amber OUT</span> = early exit &nbsp;
	&mdash; = no checkout
</div>
"""


# ── attendance Excel ─────────────────────────────────────────────────────────

def _build_attendance_excel(pe, data: dict) -> bytes:
	import openpyxl
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	dates = data["period_dates"]
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Attendance Logs"
	n_cols = 1 + len(dates)  # label column + one per day

	green = PatternFill("solid", fgColor=BRAND_GREEN)
	black = PatternFill("solid", fgColor=BRAND_BLACK)
	absent_fill = PatternFill("solid", fgColor="FBE4E4")
	white_bold = Font(bold=True, color="FFFFFF")
	center = Alignment(horizontal="center", vertical="center")
	thin = Side(style="thin", color="CCCCCC")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	# Title band
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
	c = ws.cell(row=1, column=1, value=f"{pe.company} — Attendance Logs")
	c.font = Font(bold=True, size=14, color="FFFFFF")
	c.fill = black
	c.alignment = center
	for col in range(2, n_cols + 1):
		ws.cell(row=1, column=col).fill = black
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
	c = ws.cell(row=2, column=1, value=f"{pe.start_date} to {pe.end_date}  ({pe.name})")
	c.font = Font(bold=True, color=BRAND_BLACK)
	c.fill = green
	c.alignment = center
	for col in range(2, n_cols + 1):
		ws.cell(row=2, column=col).fill = green

	row = 4
	for emp in data["employees"]:
		# Employee header (merged, black band)
		ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
		sub = " · ".join(x for x in (emp["designation"], emp["department"]) if x)
		hc = ws.cell(row=row, column=1, value=f"[{emp['code']}]  {emp['name']}  —  {sub}")
		hc.font = white_bold
		hc.fill = black
		for col in range(2, n_cols + 1):
			ws.cell(row=row, column=col).fill = black
		row += 1

		# Day-number header
		day_row = row
		lc = ws.cell(row=row, column=1, value="Date")
		lc.font = Font(bold=True, color=BRAND_BLACK)
		lc.fill = green
		lc.border = border
		for i, d in enumerate(dates, start=2):
			cell = ws.cell(row=row, column=i, value=d.day)
			cell.font = Font(bold=True, color=BRAND_BLACK, size=9)
			cell.fill = green
			cell.alignment = center
			cell.border = border
		row += 1

		# IN and OUT rows
		for label, key in (("IN", "in"), ("OUT", "out")):
			lc = ws.cell(row=row, column=1, value=label)
			lc.font = white_bold
			lc.fill = black
			lc.alignment = center
			lc.border = border
			for i, d in enumerate(dates, start=2):
				punch = data["punches"].get((emp["emp"], str(d)))
				value = ""
				cell = ws.cell(row=row, column=i)
				if punch:
					if punch["status"] == "Absent":
						value = "A" if label == "IN" else ""
						cell.fill = absent_fill
					else:
						value = punch[key]
				cell.value = value
				cell.alignment = center
				cell.border = border
				cell.font = Font(size=9)
			row += 1
		row += 1  # blank spacer between employees

	# Column widths
	ws.column_dimensions["A"].width = 12
	for i in range(2, n_cols + 1):
		ws.column_dimensions[get_column_letter(i)].width = 7
	ws.freeze_panes = "B4"

	buf = BytesIO()
	wb.save(buf)
	return buf.getvalue()
