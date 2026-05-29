"""
payroll_handler.py — Monthly Payroll Entry creation + email export (Req #3).

`create_monthly_payroll_entry()` runs every day at midnight; it fires only
on the configured Payroll Day of Month (default 30, or the last day of the
month if the month is shorter).

The Payroll Entry is created, salary slips queued, and a one-sheet Excel
file is built and attached to an email sent to the boss.
"""

from __future__ import annotations

from datetime import date
from io import BytesIO

import frappe
from frappe.utils import (
	add_months,
	flt,
	get_first_day,
	get_last_day,
	getdate,
	now_datetime,
)

from fours_customizations.fours_customizations.doctype.four_s_industries_settings.four_s_industries_settings import (
	get_settings,
)
from fours_customizations.notifications import send_email


_OFFSET_DAYS = {
	"1st day of next month": 1,
	"2nd day of next month": 2,
	"3rd day of next month": 3,
	"4th day of next month": 4,
}


def _payroll_period(settings, run_date: date) -> tuple[date, date]:
	"""Given a *run_date* (i.e. today, the day the dispatcher fires), return
	the (start, end) of the salary period this Payroll Entry should cover.

	    Run day                                Period covered
	    "Last day of payroll period month"     Same month as run_date
	    "Nth day of next month"                Previous month
	"""
	option = (settings.payroll_day_of_month or "1st day of next month").strip()
	if option == "Last day of payroll period month":
		return get_first_day(run_date), get_last_day(run_date)
	prev = add_months(run_date, -1)
	return get_first_day(prev), get_last_day(prev)


def _should_run_today(settings, today: date) -> bool:
	option = (settings.payroll_day_of_month or "1st day of next month").strip()

	if option == "Last day of payroll period month":
		return today == get_last_day(today)

	wanted_day = _OFFSET_DAYS.get(option, 1)
	return today.day == wanted_day


def daily_payroll_dispatcher() -> dict:
	settings = get_settings()
	if not int(settings.enable_payroll_automation or 0):
		return {"skipped": True}

	today = getdate(now_datetime())
	if not _should_run_today(settings, today):
		return {"skipped": True, "reason": "not the day", "today": str(today)}

	start, end = _payroll_period(settings, today)
	key = f"4s_payroll:{start.year}-{start.month}"
	if frappe.cache().get_value(key):
		return {"skipped": True, "reason": "already ran"}

	result = create_monthly_payroll_entry(settings, today, start, end)
	frappe.cache().set_value(key, "1", expires_in_sec=60 * 60 * 24 * 7)
	return result


def create_monthly_payroll_entry(settings=None, run_date=None, start=None, end=None) -> dict:
	if settings is None:
		settings = get_settings()
	if run_date is None:
		run_date = getdate(now_datetime())
	if start is None or end is None:
		start, end = _payroll_period(settings, run_date)

	company = settings.default_payroll_company or settings.default_company
	if not company:
		frappe.log_error("4S Payroll: no company configured", "4S Payroll")
		return {"error": "no company"}

	# Group employees by the Payroll Payable Account on their latest,
	# period-effective Salary Structure Assignment, then create one Payroll
	# Entry per payable account. ERPNext requires a single payable account per
	# Payroll Entry (it is mandatory), so a company whose employees book to
	# several payable accounts needs one entry each.
	emp_to_account = _employee_payable_accounts(company, end)
	if not emp_to_account:
		frappe.log_error(
			f"4S Payroll: no submitted Salary Structure Assignments with a "
			f"Payroll Payable Account for {company} effective on/before {end}",
			"4S Payroll",
		)
		return {"error": "no payable accounts"}

	created, skipped = [], []
	for account in sorted(set(emp_to_account.values())):
		if frappe.db.exists(
			"Payroll Entry",
			{
				"company": company,
				"start_date": start,
				"end_date": end,
				"payroll_payable_account": account,
				"docstatus": ("!=", 2),
			},
		):
			skipped.append(account)
			continue

		pe_name = _create_payroll_entry_for_account(
			company, run_date, start, end, account, emp_to_account
		)
		if pe_name:
			created.append(pe_name)

	if not created:
		return {"created": [], "skipped_existing": skipped}

	excel = build_payroll_excel(created)
	_send_payroll_email(settings, created, start, end, excel)

	return {"created": created, "skipped_existing": skipped}


def _employee_payable_accounts(company, period_end) -> dict:
	"""Map each employee to the Payroll Payable Account on their latest
	period-effective Salary Structure Assignment for *company*.

	Only submitted assignments dated on/before *period_end* count; when an
	employee has several, the one with the most recent ``from_date`` wins.
	Assignments without a payable account are ignored.
	"""
	rows = frappe.get_all(
		"Salary Structure Assignment",
		filters={
			"company": company,
			"docstatus": 1,
			"from_date": ["<=", period_end],
		},
		fields=["employee", "payroll_payable_account", "from_date"],
		order_by="from_date asc",
	)

	emp_to_account: dict[str, str] = {}
	for row in rows:
		if not row.payroll_payable_account:
			continue
		# Ascending order → later assignments overwrite earlier ones (latest wins).
		emp_to_account[row.employee] = row.payroll_payable_account
	return emp_to_account


def _create_payroll_entry_for_account(company, run_date, start, end, account, emp_to_account):
	"""Create, populate and submit one Payroll Entry for a single payable account.

	Employees are pulled through ERPNext's standard eligibility logic, then
	narrowed to those whose effective assignment books to *account*. Returns the
	Payroll Entry name, or ``None`` when no eligible employee belongs to it.
	"""
	pe = frappe.new_doc("Payroll Entry")
	pe.company = company
	pe.posting_date = run_date
	pe.payroll_frequency = "Monthly"
	pe.start_date = start
	pe.end_date = end
	# Set before filling so ERPNext's set_payroll_payable_account() leaves it as-is.
	pe.payroll_payable_account = account
	pe.exchange_rate = 1
	pe.flags.ignore_permissions = True

	try:
		pe.fill_employee_details()
	except Exception:
		# Older Frappe versions
		try:
			pe.get_emp_list()
		except Exception:
			frappe.log_error(frappe.get_traceback(), "4S Payroll: fill employees failed")

	# Keep only the employees that book to this payable account.
	keep = [row for row in (pe.employees or []) if emp_to_account.get(row.employee) == account]
	if not keep:
		return None
	pe.set("employees", keep)
	pe.number_of_employees = len(keep)
	# Re-assert in case fill_employee_details adjusted it.
	pe.payroll_payable_account = account

	pe.insert(ignore_permissions=True)

	# Creating + submitting salary slips can fail per-run; don't let it abort the
	# remaining accounts or the summary email.
	try:
		pe.submit()
	except Exception:
		frappe.log_error(frappe.get_traceback(), f"4S Payroll: submit failed for {pe.name} (continuing)")

	return pe.name


def build_payroll_excel(payroll_entries) -> bytes:
	"""Build a one-sheet Excel file with one row per Salary Slip across the
	given Payroll Entry (or list of Payroll Entries)."""
	import openpyxl

	if isinstance(payroll_entries, str):
		payroll_entries = [payroll_entries]

	slips = frappe.get_all(
		"Salary Slip",
		filters={"payroll_entry": ["in", payroll_entries]},
		fields=[
			"name", "employee", "employee_name", "department", "designation",
			"gross_pay", "total_deduction", "net_pay", "bank_account_no", "bank_name",
			"payroll_entry",
		],
		order_by="employee_name",
	)

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Payroll"

	headers = [
		"Employee", "Employee Name", "Department", "Designation",
		"Bank", "Account No.",
		"Gross Pay", "Total Deductions", "Net Pay",
		"Payroll Entry",
	]
	ws.append(headers)

	for h in ws[1]:
		h.font = openpyxl.styles.Font(bold=True)

	total_gross = total_deductions = total_net = 0.0

	for slip in slips:
		ws.append([
			slip.employee,
			slip.employee_name,
			slip.department or "",
			slip.designation or "",
			slip.bank_name or "",
			slip.bank_account_no or "",
			flt(slip.gross_pay),
			flt(slip.total_deduction),
			flt(slip.net_pay),
			slip.payroll_entry or "",
		])
		total_gross += flt(slip.gross_pay)
		total_deductions += flt(slip.total_deduction)
		total_net += flt(slip.net_pay)

	ws.append([])
	ws.append(["TOTAL", "", "", "", "", "", total_gross, total_deductions, total_net, ""])
	for c in ws[ws.max_row]:
		c.font = openpyxl.styles.Font(bold=True)

	# Autosize columns
	for column in ws.columns:
		length = max((len(str(cell.value)) for cell in column if cell.value is not None), default=10)
		col_letter = column[0].column_letter
		ws.column_dimensions[col_letter].width = min(length + 2, 40)

	buf = BytesIO()
	wb.save(buf)
	return buf.getvalue()


def _send_payroll_email(settings, payroll_entries, start, end, excel_bytes: bytes) -> None:
	to = settings.boss_email
	if not to:
		frappe.log_error("4S Payroll: boss_email not configured", "4S Payroll")
		return
	if isinstance(payroll_entries, str):
		payroll_entries = [payroll_entries]
	subject = f"Payroll Run — {end}"
	pe_items = "".join(f"<li><b>{name}</b></li>" for name in payroll_entries)
	body = f"""
<p>Dear Sir / Madam,</p>
<p>Attached is the payroll run for the period <b>{start}</b> to <b>{end}</b>.</p>
<p>The following Payroll Entries were created (one per Payroll Payable Account):</p>
<ul>
  {pe_items}
</ul>
<p>Please review and approve in ERPNext.</p>
"""
	filename = f"payroll-{end}.xlsx"
	send_email(
		subject=subject,
		message=body,
		recipients=to,
		cc=settings.payroll_cc_emails,
		attachments=[{"fname": filename, "fcontent": excel_bytes}],
	)
